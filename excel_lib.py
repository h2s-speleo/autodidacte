#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jun  3 14:34:51 2020

@author: j
"""
import openpyxl

class Tableur():
    """
    classe Tableur
    
    parametres :
        nom - obligatoire   nom du fichier
        path - défaut ""   chemin vers le fichier
        nomFe - defaut None   nom de la feuille
        nouv - créer un nouveau fichier : liste des champs
        
    autres attributs :
        tableur = objet openpyxl.workbook
        feActive - objet apenpyxl.worksheet
        index - dictionnaire['nom du champ'] = numéro de colonne
    """
    def __init__(self, **kwargs):
        if 'nom' in kwargs.keys():
            self.nom = kwargs.get('nom')
        else :
            raise AttributeError("EREURE : l'atribut nom est obligatoire"
                                 + " (nom du fichier xlsx)")
        if 'path' in kwargs.keys():
            self.path = kwargs.get('path')
        else :
            self.path = ''
        if 'nomFe' in kwargs.keys():
            self.nomFe = kwargs.get('nomFe')
        else :
            self.nomFe = None
        if 'nouv' in kwargs.keys():
            self.create(kwargs.get('nouv'))
        else :
            self.open()
            
            
    def open(self):
        """
        ouverture d'un fichier existant
        """
        self.tableur = openpyxl.load_workbook(filename = 
                                              str(self.path + self.nom))
#        print('Chargement du fichier {}\n'
#              'chemin : {}'
#              .format(self.nom, self.path))
        if len(self.tableur.sheetnames) == 1 :
            self.nomFe = self.tableur.sheetnames[0]
#            print('une seule feuille disponible')
        elif self.nomFe not in self.tableur.sheetnames :
            self.choixFeuille()
#        print('chargement de la feuille "{}"'
#              .format(self.nomFe))
        self.feActive = self.tableur[self.nomFe]
        self.creatIndex()
#        print()
        
    def create(self, arg):
        """création d'un nouveau fichier.
        arg : liste des champs
        """
        self.tableur = openpyxl.Workbook()
        self.nomFe = self.tableur.sheetnames[0]
        self.feActive = self.tableur[self.nomFe]
        self.feActive.append(arg)
        self.creatIndex()
        self.save()
        
    def creatIndex(self):
        """
        création de l'index de la feuille
        """
        self.index = dict()
        for i in range(self.feActive.max_column):
            self.index[self.feActive.cell(1,i+1).value] = i+1
#        print(self.index)
#        print()

    def choixFeuille(self):
        """
        choix de la feuille dans le tableur
        """
        listeF=self.tableur.sheetnames
        while True :
            print("\nchoix d'une feuille dans le tableau :")
            for i in range(len(listeF)):
                print('{} : "{}"'
                      .format(i, listeF[i]))
            print("choix de la feuille :")
            x = input()
            try :
                x = int(x)
                if x in range(len(listeF)):
                    self.nomFe = listeF[i]
                    break
                else :
                    print('"{}" ne fait pas partie des options proposées'
                          .format(x))
            except:
                print("la réponse doit etre un nombre")
        print()
        
    def ajoutLigne(self, dico):
        """
        ajoute une ligne en fin de tableau avec les valeurs donnée dans 
        le dico
        """
        x = self.feActive.max_row + 1
        self.ajoutValeur(x, dico)
        
    def ajoutValeur(self, ligne, dico):
        """
        ajoute les valeurs du dico dans la ligne spécifiée
        """
        for key in dico :
            y = self.index[key]
#            print('coordonés : ligne = {} collone = {}. champ "{}" - '
#                  'valeur = "{}"'
#                  .format(ligne, y, key, dico[key]))
            self.feActive.cell(row = ligne, column = y).value = dico[key]
#        print()
        
    def save(self):
        """
        enregistrer le tableau
        """
        self.tableur.save(filename = str(self.path + self.nom))
        
    
    def saveAs(self, **kwargs):
        """
        enregistrer le dico sous un autre nom
        kwargs :
            - nom : nom du fichier (obligatoire, inclus ".xlsx")
            - path : chemin vers le fichier, défaut ""
        """
        if 'nom' in kwargs.keys():
            nom = kwargs.get('nom')
        else :
            raise AttributeError("EREURE : l'atribut nom est obligatoire"
                                 + " (nom du fichier xlsx)")
        if 'path' in kwargs.keys():
            path = kwargs.get('path')
        else :
            path = self.path
        
        self.tableur.save(filename = str(path + nom))
        
    def rechercheValeur(self,*args):
        """
        générateur, renvoie un dico pour chaque ligne
        les clés du dictionnaire sont les valeurs passés dans args (liste)
        """
        if len(args) == 0:
            temp = list()
            for key in self.index:
                temp.append(key)
            args = tuple(temp)
        for row in self.feActive.iter_rows(min_row = 2) :
            dico = {}
            for i in range(len(args)):
                dico[args[i]] = row[self.index[args[i]]-1].value
            yield (dico)
            
    def rechercheLigne(self,*args):
        """
        recherche l'indexe de toutes las ligne qui respectent toutes les
        égalitées passées dans la liste args
        (('nom champ1', 'val champ1), ('nom champ2', 'val champ2))
        
        exemple :
            test.rechercheLigne( ('nom', 'A'), ('ID', 1) )
        retourne une liste d'entiers
        """
        listeLigne = list()
        for i in range(1,self.feActive.max_row):
            vraix = 0
            for j in range(len(args)):
                if self.feActive.cell(i+1,
                                      self.index[args[j][0]]
                                      ).value == args[j][1] :
                    vraix +=1
            if vraix == len(args):
                listeLigne.append(i+1)
        return listeLigne
    
    def afficherLigne(self, arg, *args):
        """
        renvoie les valeurs dans la ligne spécifiée, si aucun champ n'est
        spécifiée dans la liste args, tout les champs sont sont renvoyé
        retourne un dictionnaire['nom champ'] = 'valeur'
        """
        args = list(args)
        dico = dict()
        
        if len(args) == 0:
            for key in self.index:
                args.append(key)
        for i in range(len(args)) :
            dico[args[i]] = self.feActive.cell(arg,
                                               self.index[args[i]]
                                               ).value
        return dico
        

if __name__ == "__main__":
    
    # création du fichier test
    test = Tableur(nom = 'test.xlsx',
                    path = '',
                    nouv = ['ID', 'nom', 'val'])
    
    DATA = [
            {'ID': 1, 'nom': 'A', 'val': 22},
            {'ID': 2, 'nom': 'B', 'val': 5},
            {'ID': 3, 'nom': 'C', 'val': 86},
            {'ID': 4, 'nom': 'D', 'val': 44},
            {'ID': 5, 'nom': 'E', 'val': 69}
            ]
    for i in range(len(DATA)):
        test.ajoutLigne(DATA[i])
    test.save()
    del(test)
    

    # ouverture du tableau
    test = Tableur(nom = "test.xlsx", nomFe = "testSheet")
    print ()
    
    # recherche de valeur dans tout le tableau (génératreur)
    for i in test.rechercheValeur('ID', 'nom', 'val'):
        print(i)
    print()
    
    # recherche l'indexe de la ligne qui remplis toute les conditions
    print(  test.rechercheLigne( ('nom', 'A'), ('ID', 1))  )
    
    # ajout d'une ligne dans le tableau
    dico = {'ID': 6, 'nom': 'F', 'val': 666}
    test.ajoutLigne(dico)
    
    # changement de la valeur dans la ligne 7
    dico2 = {'val' : 555}
    test.ajoutValeur(7, dico2)
    
    # supprimer une ligne (ici la dernière)
    test.feActive.delete_rows(test.feActive.max_row)
    
    # enregistrement du tableau
    test.save()
    
    # affiche les valeur des champs de la ligne indiquée
    print(   test.afficherLigne(2, 'nom','val')   )
    
